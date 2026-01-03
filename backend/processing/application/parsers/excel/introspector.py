import io

from openpyxl import load_workbook
from processing.application.parsers.excel.contracts import (
    ExcelIntrospection,
    SheetIntrospection,
)


class ExcelIntrospector:
    """
    Niveau 0 — Introspection
    AUCUNE interprétation métier
    Compatible fichiers Excel très lourds
    """

    @staticmethod
    def inspect(content: bytes) -> ExcelIntrospection:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

        sheets_meta = []

        for sheet in wb.worksheets:
            rows = sheet.max_row or 0
            cols = sheet.max_column or 0
            total_cells = max(rows * cols, 1)

            non_empty = 0
            empty_top = 0
            empty_bottom = 0

            # --- empty rows top
            for row in sheet.iter_rows():
                if all(c.value in (None, "") for c in row):
                    empty_top += 1
                else:
                    break

            # --- empty rows bottom (SAFE)
            all_rows = list(sheet.iter_rows())
            for row in reversed(all_rows):
                if all(c.value in (None, "") for c in row):
                    empty_bottom += 1
                else:
                    break

            # --- non empty count
            for row in all_rows:
                for cell in row:
                    if cell.value not in (None, ""):
                        non_empty += 1

            sheets_meta.append(
                SheetIntrospection(
                    name=sheet.title,
                    rows=rows,
                    columns=cols,
                    non_empty_ratio=non_empty / total_cells,
                    merged_cells_count=-1,  # 🔥 inconnu en read_only
                    empty_rows_top=empty_top,
                    empty_rows_bottom=empty_bottom,
                )
            )

        return ExcelIntrospection(
            file_size=len(content),
            sheets=sheets_meta,
        )
